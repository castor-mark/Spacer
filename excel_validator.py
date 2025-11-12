import os
import re
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
import shutil
from config import EXCEL_FOLDER, OUTPUT_FOLDER

def find_spacing_issues(text, strict_mode=False):
    """
    Check for spacing issues and special characters (excluding ._:-)
    
    Args:
        text: The text to check
        strict_mode: If True, ANY space is flagged (for columns that should never have spaces)
    
    Returns tuple: (has_issues, issue_description, suggested_fix)
    """
    if not isinstance(text, str):
        return False, "", text
    
    issues = []
    suggested_fix = text
    
    # STRICT MODE: Flag ANY spaces at all
    if strict_mode:
        if " " in text:
            space_count = text.count(" ")
            issues.append(f"contains {space_count} space(s)")
            suggested_fix = text.replace(' ', '')  # Remove ALL spaces
        return len(issues) > 0, " | ".join(issues), suggested_fix
    
    # NORMAL MODE: Only flag problematic spaces
    # Check for double spaces
    if "  " in text:
        issues.append("double spaces")
        suggested_fix = re.sub(r'\s{2,}', ' ', suggested_fix)
    
    # Check for leading/trailing spaces
    if text != text.strip():
        issues.append("leading/trailing spaces")
        suggested_fix = suggested_fix.strip()
    
    # Check for spaces before underscores, periods, hyphens, or colons
    if re.search(r' [._:-]', text):
        issues.append("space before ._:-")
        suggested_fix = re.sub(r' ([._:-])', r'\1', suggested_fix)
    
    # Check for spaces after underscores, periods, hyphens, or colons
    if re.search(r'[._:-] ', text):
        issues.append("space after ._:-")
        suggested_fix = re.sub(r'([._:-]) ', r'\1', suggested_fix)
    
    # Check for special characters (excluding letters, numbers, spaces, ._:-)
    special_chars = re.findall(r'[^\w\s._:-]', text)
    if special_chars:
        unique_chars = list(set(special_chars))
        issues.append(f"special characters: {''.join(unique_chars)}")
    
    # Check for multiple consecutive spaces (3 or more)
    if re.search(r'\s{3,}', text):
        issues.append("multiple consecutive spaces")
        suggested_fix = re.sub(r'\s{3,}', ' ', suggested_fix)
    
    return len(issues) > 0, " | ".join(issues), suggested_fix

def validate_24hour_format(text):
    """
    Check if datetime string uses proper 24-hour format with leading zeros
    Returns tuple: (has_issues, issue_description, suggested_fix)
    
    Valid format: YYYY-MM-DDTHH:MM:SS (hours must be 00-23 with leading zero)
    """
    if not isinstance(text, str):
        return False, "", text
    
    text = text.strip()
    
    # Pattern for datetime with potential time issues
    datetime_pattern = r'(\d{4}[-/]\d{2}[-/]\d{2})[T\s](\d{1,2}):(\d{2}):(\d{2})'
    
    match = re.search(datetime_pattern, text)
    
    if not match:
        # Not a datetime format we're checking
        return False, "", text
    
    date_part = match.group(1)
    hour = match.group(2)
    minute = match.group(3)
    second = match.group(4)
    
    issues = []
    suggested_fix = text
    
    # Check if hour has leading zero (should be 2 digits)
    if len(hour) == 1:
        issues.append(f"Hour missing leading zero: '{hour}' should be '0{hour}'")
        old_time = f"{hour}:{minute}:{second}"
        new_time = f"0{hour}:{minute}:{second}"
        suggested_fix = text.replace(old_time, new_time)
    elif len(hour) == 2 and hour[0] == '0':
        pass
    elif int(hour) > 23:
        issues.append(f"Invalid hour: '{hour}' (must be 00-23)")
    
    # Check minute format
    if len(minute) == 1:
        issues.append(f"Minute missing leading zero: '{minute}' should be '0{minute}'")
        old_time = f"{hour}:{minute}:{second}"
        new_hour = f"0{hour}" if len(hour) == 1 else hour
        new_time = f"{new_hour}:0{minute}:{second}"
        suggested_fix = text.replace(old_time, new_time)
    
    # Check second format
    if len(second) == 1:
        issues.append(f"Second missing leading zero: '{second}' should be '0{second}'")
    
    return len(issues) > 0, " | ".join(issues), suggested_fix

def validate_file_extension_case(text):
    """
    Check if file extensions are in lowercase (case-sensitive check)
    Returns tuple: (has_issues, issue_description, suggested_fix)
    
    Examples:
    - file.PDF -> file.pdf (has issue)
    - file.Pdf -> file.pdf (has issue)
    - file.pdf -> file.pdf (no issue)
    - file.XLSX -> file.xlsx (has issue)
    """
    if not isinstance(text, str):
        return False, "", text
    
    text = text.strip()
    
    # Common file extensions to check (add more as needed)
    common_extensions = [
        'pdf', 'xlsx', 'xls', 'doc', 'docx', 'ppt', 'pptx',
        'txt', 'csv', 'json', 'xml', 'html', 'htm',
        'jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp',
        'mp3', 'mp4', 'avi', 'mov', 'wmv', 'flv', 'mkv',
        'zip', 'rar', '7z', 'tar', 'gz', 'bz2',
        'py', 'js', 'java', 'cpp', 'c', 'h', 'css', 'php', 'rb', 'go',
        'md', 'rst', 'tex', 'log', 'dat', 'sql', 'db',
        'exe', 'dll', 'so', 'app', 'deb', 'rpm',
        'ai', 'psd', 'eps', 'indd', 'sketch',
        'mp3', 'wav', 'flac', 'aac', 'ogg', 'wma',
        'eml', 'msg', 'ics', 'vcf'
    ]
    
    # Pattern to find file extensions (handles multiple dots in filename)
    pattern = r'\.([a-zA-Z0-9]+)$'
    
    match = re.search(pattern, text)
    
    if not match:
        # No file extension found
        return False, "", text
    
    extension = match.group(1)
    extension_lower = extension.lower()
    
    # Check if the extension is not lowercase
    if extension != extension_lower:
        # Extension has uppercase letters
        issues = f"Extension '.{extension}' should be lowercase '.{extension_lower}'"
        suggested_fix = text[:match.start()] + '.' + extension_lower
        return True, issues, suggested_fix
    
    return False, "", text

def list_excel_files():
    """
    List all Excel files in the EXCEL_FOLDER
    """
    excel_files = []
    for file in os.listdir(EXCEL_FOLDER):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(file)
    
    return excel_files

def list_sheet_names(file_path):
    """
    List all sheet names in an Excel file
    """
    try:
        xl_file = pd.ExcelFile(file_path)
        return xl_file.sheet_names
    except Exception as e:
        print(f"Error reading sheets: {e}")
        return []

def list_available_columns(file_path, sheet_name):
    """
    List all available column headers in the specified sheet
    Returns a list of (index, name) tuples
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
        headers = list(df.columns)
        clean_headers = [(i+1, str(h).strip()) for i, h in enumerate(headers) if str(h).strip()]
        
        print(f"\n📋 Available columns in '{sheet_name}':")
        for index, header in clean_headers:
            print(f"  {index}. {header}")
        return clean_headers
    except Exception as e:
        print(f"❌ Error listing columns: {e}")
        return []

def get_user_inputs():
    """
    Get file name, sheet name, column names, and validation type from user input
    """
    print("🔍 ENHANCED EXCEL VALIDATOR")
    print("=" * 50)
    print("This script validates:")
    print("  📝 Spacing issues and special characters")
    print("  ⏰ 24-hour time format (e.g., 05:11:20 not 5:11:20)")
    print("  📁 File extension case (e.g., .pdf not .PDF)")
    print("=" * 50)
    
    # List available Excel files
    excel_files = list_excel_files()
    if not excel_files:
        print(f"❌ No Excel files found in '{EXCEL_FOLDER}' folder!")
        return None, None, None, None, None
    
    print(f"\n📋 Available Excel files in '{EXCEL_FOLDER}' folder:")
    for i, file in enumerate(excel_files, 1):
        print(f"  {i}. {file}")
    
    file_choice = input("\nEnter the number of the file to process: ").strip()
    try:
        file_index = int(file_choice) - 1
        if file_index < 0 or file_index >= len(excel_files):
            print("❌ Invalid file number!")
            return None, None, None, None, None
        excel_file = excel_files[file_index]
    except ValueError:
        print("❌ Please enter a valid number!")
        return None, None, None, None, None
    
    file_path = os.path.join(EXCEL_FOLDER, excel_file)
    
    # List available sheets
    sheet_names = list_sheet_names(file_path)
    if not sheet_names:
        print(f"❌ No sheets found in '{excel_file}'!")
        return None, None, None, None, None
    
    print(f"\n📋 Available sheets in '{excel_file}':")
    for i, sheet in enumerate(sheet_names, 1):
        print(f"  {i}. {sheet}")
    
    sheet_choice = input("\nEnter the number of the sheet to process: ").strip()
    try:
        sheet_index = int(sheet_choice) - 1
        if sheet_index < 0 or sheet_index >= len(sheet_names):
            print("❌ Invalid sheet number!")
            return None, None, None, None, None
        sheet_name = sheet_names[sheet_index]
    except ValueError:
        print("❌ Please enter a valid number!")
        return None, None, None, None, None
    
    # List available columns
    available_columns = list_available_columns(file_path, sheet_name)
    if not available_columns:
        print(f"❌ No columns found in '{sheet_name}'!")
        return None, None, None, None, None
    
    # Choose validation type
    print("\nValidation Type:")
    print("1. Spacing & Special Characters only")
    print("2. 24-Hour Time Format only")
    print("3. File Extension Case only")
    print("4. Spacing & Time Format")
    print("5. Spacing & File Extension")
    print("6. Time Format & File Extension")
    print("7. All validations")
    
    validation_choice = input("Choose option (1-7): ").strip()
    
    if validation_choice not in ['1', '2', '3', '4', '5', '6', '7']:
        print("❌ Invalid choice!")
        return None, None, None, None, None
    
    validation_types = {
        '1': ['spacing'],
        '2': ['time'],
        '3': ['extension'],
        '4': ['spacing', 'time'],
        '5': ['spacing', 'extension'],
        '6': ['time', 'extension'],
        '7': ['spacing', 'time', 'extension']
    }
    
    validations = validation_types[validation_choice]
    
    # Get column names by number
    print("\nColumn Selection Options:")
    print("1. Single column (by number)")
    print("2. Multiple columns (comma-separated numbers)")
    print("3. All columns")
    
    choice = input("Choose option (1-3): ").strip()
    
    if choice == "1":
        column_choice = input("Enter column number: ").strip()
        try:
            column_index = int(column_choice) - 1
            if column_index < 0 or column_index >= len(available_columns):
                print("❌ Invalid column number!")
                return None, None, None, None, None
            column_names = [available_columns[column_index][1]]
        except ValueError:
            print("❌ Please enter a valid number!")
            return None, None, None, None, None
    elif choice == "2":
        columns_input = input("Enter column numbers separated by commas: ").strip()
        if not columns_input:
            print("❌ Column numbers are required!")
            return None, None, None, None, None
        
        try:
            column_indices = [int(num.strip()) - 1 for num in columns_input.split(',')]
            column_names = []
            for index in column_indices:
                if index < 0 or index >= len(available_columns):
                    print(f"❌ Invalid column number: {index + 1}")
                    return None, None, None, None, None
                column_names.append(available_columns[index][1])
        except ValueError:
            print("❌ Please enter valid numbers!")
            return None, None, None, None, None
    elif choice == "3":
        column_names = ["ALL_COLUMNS"]
    else:
        print("❌ Invalid choice!")
        return None, None, None, None, None
    
    # Ask about strict mode for spacing validation
    strict_columns = []
    if 'spacing' in validations and column_names != ["ALL_COLUMNS"]:
        print("\n🔴 STRICT MODE for Spacing:")
        print("   In strict mode, ANY space (including single spaces between words) is flagged.")
        print("   This is useful for columns like filenames, IDs, SKUs that should NEVER contain spaces.")
        print()
        print(f"Selected columns: {', '.join(column_names)}")
        print()
        print("Options:")
        print("1. Apply STRICT mode to ALL selected columns")
        print("2. Apply STRICT mode to SOME columns (specify by number)")
        print("3. Use NORMAL mode for all columns (only problematic spaces flagged)")
        
        strict_choice = input("\nChoose option (1-3): ").strip()
        
        if strict_choice == "1":
            strict_columns = column_names.copy()
            print(f"\n✅ Strict mode will be applied to ALL selected columns: {', '.join(strict_columns)}")
        elif strict_choice == "2":
            print("\nAvailable columns for strict mode:")
            for index, header in available_columns:
                print(f"  {index}. {header}")
            
            strict_input = input("Enter column numbers for STRICT mode (comma-separated): ").strip()
            if strict_input:
                try:
                    strict_indices = [int(num.strip()) - 1 for num in strict_input.split(',')]
                    strict_columns = []
                    for index in strict_indices:
                        if index < 0 or index >= len(available_columns):
                            print(f"❌ Invalid column number: {index + 1}")
                            continue
                        strict_columns.append(available_columns[index][1])
                    print(f"\n✅ Strict mode will be applied to: {', '.join(strict_columns)}")
                except ValueError:
                    print("❌ Please enter valid numbers!")
            else:
                print("\n✅ No columns in strict mode. All spacing checks will use normal mode.")
        else:
            print("\n✅ Using NORMAL mode for all columns (only problematic spaces flagged).")
    elif 'spacing' in validations and column_names == ["ALL_COLUMNS"]:
        print("\n🔴 STRICT MODE for Spacing:")
        print("   You're checking ALL columns. Enter specific column numbers that should use strict mode.")
        print("   (In strict mode, ANY space is flagged - useful for filenames, IDs, SKUs)")
        print()
        print("Available columns for strict mode:")
        for index, header in available_columns:
            print(f"  {index}. {header}")
        
        strict_input = input("Enter column numbers for STRICT mode (comma-separated, or press Enter to skip): ").strip()
        
        if strict_input:
            try:
                strict_indices = [int(num.strip()) - 1 for num in strict_input.split(',')]
                strict_columns = []
                for index in strict_indices:
                    if index < 0 or index >= len(available_columns):
                        print(f"❌ Invalid column number: {index + 1}")
                        continue
                    strict_columns.append(available_columns[index][1])
                print(f"\n✅ Strict mode will be applied to: {', '.join(strict_columns)}")
            except ValueError:
                print("❌ Please enter valid numbers!")
        else:
            print("\n✅ All columns will use NORMAL mode (only problematic spaces flagged).")
    
    return file_path, sheet_name, column_names, validations, strict_columns

def check_columns(file_path, sheet_name, column_headers, validations, strict_columns=None):
    """
    Check multiple columns for various issues based on selected validations
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet/tab
        column_headers: List of column names to check, or ["ALL_COLUMNS"]
        validations: List of validation types ['spacing', 'time', 'extension']
        strict_columns: List of column names where ANY space should be flagged
    """
    if strict_columns is None:
        strict_columns = []
    
    if column_headers == ["ALL_COLUMNS"]:
        print(f"\n🔍 Analyzing ALL columns in sheet '{sheet_name}'...")
    else:
        print(f"\n🔍 Analyzing {len(column_headers)} columns in sheet '{sheet_name}'...")
        print(f"Columns: {', '.join(column_headers)}")
    
    print(f"Validation types: {', '.join(validations)}")
    
    if strict_columns:
        print(f"🔴 Strict mode columns (ANY space flagged): {', '.join(strict_columns)}")
    
    print("=" * 60)
    
    try:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        print(f"✅ Opened file: {os.path.basename(file_path)} - {sheet_name}")
        print(f"Found {len(df.columns)} columns and {len(df)} rows")
        
        # Get and analyze headers
        headers = list(df.columns)
        
        # Determine which columns to check
        if column_headers == ["ALL_COLUMNS"]:
            target_columns = []
            for i, header in enumerate(headers):
                if str(header).strip():
                    target_columns.append((i, str(header).strip()))
        else:
            target_columns = []
            for column_header in column_headers:
                found = False
                for i, header in enumerate(headers):
                    if str(header).strip().lower() == column_header.lower():
                        target_columns.append((i, str(header).strip()))
                        found = True
                        break
                
                if not found:
                    print(f"\n⚠️ Warning: Column '{column_header}' not found!")
                    available_headers = [str(h).strip() for h in headers if str(h).strip()]
                    suggestions = [h for h in available_headers if column_header.lower() in h.lower()]
                    if suggestions:
                        print(f"Did you mean one of these? {suggestions}")
        
        if not target_columns:
            print("❌ No valid columns found to check!")
            return None
        
        print(f"✅ Found {len(target_columns)} columns to analyze:")
        for col_index, col_name in target_columns:
            mode_label = "🔴 STRICT" if col_name in strict_columns else "🔵 NORMAL"
            print(f"  {mode_label} {col_name} (position {col_index + 1})")
        
        # Check each column for issues
        all_issues = []
        
        print(f"\n🔍 Running validations...")
        
        for col_idx, (target_column_index, column_name) in enumerate(target_columns):
            is_strict = column_name in strict_columns
            mode_label = "🔴 STRICT" if is_strict else "🔵 NORMAL"
            print(f"\n📋 Checking column '{column_name}' [{mode_label}]...")
            
            # Extract values from this column
            column_values = df.iloc[:, target_column_index].fillna("").astype(str).tolist()
            
            # Check for issues in this column
            column_issues_count = 0
            
            for i, value in enumerate(column_values):
                if not value.strip():
                    continue
                
                row_number = i + 2  # Excel rows are 1-indexed, and header is row 1
                combined_issues = []
                combined_fix = value
                
                # Run spacing validation
                if 'spacing' in validations:
                    has_spacing_issues, spacing_desc, spacing_fix = find_spacing_issues(value, strict_mode=is_strict)
                    if has_spacing_issues:
                        mode_prefix = "[Strict Spacing]" if is_strict else "[Spacing]"
                        combined_issues.append(f"{mode_prefix} {spacing_desc}")
                        combined_fix = spacing_fix
                
                # Run time format validation
                if 'time' in validations:
                    has_time_issues, time_desc, time_fix = validate_24hour_format(value)
                    if has_time_issues:
                        combined_issues.append(f"[Time Format] {time_desc}")
                        combined_fix = time_fix
                
                # Run file extension validation
                if 'extension' in validations:
                    has_ext_issues, ext_desc, ext_fix = validate_file_extension_case(value)
                    if has_ext_issues:
                        combined_issues.append(f"[File Extension] {ext_desc}")
                        combined_fix = ext_fix
                
                if combined_issues:
                    issue_data = {
                        'column': column_name,
                        'column_index': target_column_index,
                        'row': row_number,
                        'original_value': value,
                        'display_value': repr(value),
                        'issues': " | ".join(combined_issues),
                        'suggested_fix': combined_fix,
                        'is_strict': is_strict
                    }
                    all_issues.append(issue_data)
                    column_issues_count += 1
            
            # Report results for this column
            if column_issues_count > 0:
                print(f"  ❌ Found {column_issues_count} issues")
            else:
                print(f"  ✅ No issues found")
        
        # Overall results summary
        print(f"\n📊 OVERALL ANALYSIS RESULTS:")
        print("=" * 50)
        print(f"Total columns checked: {len(target_columns)}")
        print(f"Total cells with issues: {len(all_issues)}")
        
        if all_issues:
            print(f"\n⚠️  ISSUES FOUND:")
            print("-" * 40)
            
            # Group issues by column
            issues_by_column = {}
            for issue in all_issues:
                col_name = issue['column']
                if col_name not in issues_by_column:
                    issues_by_column[col_name] = []
                issues_by_column[col_name].append(issue)
            
            print("Issues by column:")
            for col_name, col_issues in issues_by_column.items():
                mode_label = "🔴 STRICT" if col_name in strict_columns else "🔵 NORMAL"
                print(f"  {mode_label} {col_name}: {len(col_issues)} issues")
            
            # Count issue types
            spacing_count = sum(1 for issue in all_issues if '[Spacing]' in issue['issues'] or '[Strict Spacing]' in issue['issues'])
            time_count = sum(1 for issue in all_issues if '[Time Format]' in issue['issues'])
            extension_count = sum(1 for issue in all_issues if '[File Extension]' in issue['issues'])
            
            print(f"\nIssues by validation type:")
            if spacing_count > 0:
                strict_spacing_count = sum(1 for issue in all_issues if '[Strict Spacing]' in issue['issues'])
                normal_spacing_count = spacing_count - strict_spacing_count
                if strict_spacing_count > 0:
                    print(f"  🔴 Strict Spacing (any space): {strict_spacing_count}")
                if normal_spacing_count > 0:
                    print(f"  🔵 Normal Spacing/Special Characters: {normal_spacing_count}")
            if time_count > 0:
                print(f"  ⏰ Time Format: {time_count}")
            if extension_count > 0:
                print(f"  📁 File Extension Case: {extension_count}")
            
            print(f"\nDetailed issues (showing first 20):")
            for i, issue in enumerate(all_issues[:20], 1):
                mode_icon = "🔴" if issue.get('is_strict') else "🔵"
                print(f"\n{i:2d}. {mode_icon} Column '{issue['column']}', Row {issue['row']:3d}")
                print(f"    Issue:   {issue['issues']}")
                print(f"    Current: {issue['display_value']}")
                print(f"    Fixed:   '{issue['suggested_fix']}'")
            
            if len(all_issues) > 20:
                print(f"\n... and {len(all_issues) - 20} more issues")
            
            print(f"\nCell locations with issues (first 30):")
            for issue in all_issues[:30]:
                mode_icon = "🔴" if issue.get('is_strict') else "🔵"
                print(f"  {mode_icon} Row {issue['row']}, Column {issue['column_index']+1} ({issue['column']})")
            if len(all_issues) > 30:
                print(f"  ... and {len(all_issues) - 30} more cells")
        
        return all_issues
    
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_report_and_highlight(file_path, sheet_name, issues, session_issues=None):
    """
    Create a detailed report and highlight issues in the Excel file
    
    Args:
        file_path: Path to the original Excel file
        sheet_name: Name of the sheet/tab
        issues: List of issues found in current analysis
        session_issues: List of all issues found in the current session (optional)
    """
    try:
        # Create a timestamp for the report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        
        # Create timestamped folder
        timestamped_folder = os.path.join(OUTPUT_FOLDER, timestamp)
        os.makedirs(timestamped_folder, exist_ok=True)
        
        # Create latest folder
        latest_folder = os.path.join(OUTPUT_FOLDER, "latest")
        os.makedirs(latest_folder, exist_ok=True)
        
        # Determine which issues to use for the report
        report_issues = session_issues if session_issues is not None else issues
        
        # Create the report filename
        report_filename = f"{base_filename}_{sheet_name}_validation_report_{timestamp}.xlsx"
        report_path = os.path.join(timestamped_folder, report_filename)
        
        # Also create a copy in the latest folder
        latest_report_path = os.path.join(latest_folder, report_filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in workbook!")
            return
        
        ws = wb[sheet_name]
        
        # Define styles for highlighting
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        white_font = Font(color="FFFFFF")
        
        # Highlight cells with issues
        for issue in issues:
            row = issue['row']
            col = issue['column_index'] + 1  # openpyxl is 1-indexed
            cell = ws.cell(row=row, column=col)
            cell.fill = red_fill
            cell.font = white_font
        
        # Create a new sheet for the report
        report_sheet_name = f"Validation Report"
        if report_sheet_name in wb.sheetnames:
            wb.remove(wb[report_sheet_name])
        report_ws = wb.create_sheet(title=report_sheet_name)
        
        # Add headers to the report
        report_ws.append(["Row", "Column", "Original Value", "Issues", "Suggested Fix"])
        
        # Add issue details to the report
        for issue in report_issues:
            report_ws.append([
                issue['row'],
                issue['column'],
                issue['original_value'],
                issue['issues'],
                issue['suggested_fix']
            ])
        
        # Save the workbook to the timestamped folder
        wb.save(report_path)
        
        # Also save a copy to the latest folder
        wb.save(latest_report_path)
        
        print(f"✅ Report created and issues highlighted!")
        print(f"📁 Saved to timestamped folder: {report_path}")
        print(f"📁 Also saved to latest folder: {latest_report_path}")
        
        # Also create a CSV report for easier viewing
        csv_filename = f"{base_filename}_{sheet_name}_validation_report_{timestamp}.csv"
        csv_path = os.path.join(timestamped_folder, csv_filename)
        latest_csv_path = os.path.join(latest_folder, csv_filename)
        
        # Convert report data to DataFrame and save as CSV
        report_data = []
        for issue in report_issues:
            report_data.append({
                "Row": issue['row'],
                "Column": issue['column'],
                "Original Value": issue['original_value'],
                "Issues": issue['issues'],
                "Suggested Fix": issue['suggested_fix']
            })
        
        df_report = pd.DataFrame(report_data)
        df_report.to_csv(csv_path, index=False)
        df_report.to_csv(latest_csv_path, index=False)
        
        print(f"📁 CSV report also saved to: {csv_path}")
        print(f"📁 CSV report also saved to latest folder: {latest_csv_path}")
    
    except Exception as e:
        print(f"❌ Error creating report: {e}")

def main():
    """
    Main interactive function
    """
    # Initialize session issues list to accumulate all issues in the session
    session_issues = []
    
    while True:
        # Get user inputs
        file_path, sheet_name, column_names, validations, strict_columns = get_user_inputs()
        
        if not all([file_path, sheet_name, column_names, validations is not None]):
            continue
        
        # Run the analysis
        issues = check_columns(file_path, sheet_name, column_names, validations, strict_columns)
        
        if issues is not None:
            # Add current issues to the session issues list
            session_issues.extend(issues)
            
            # Ask if they want to create a report
            report_choice = input(f"\n📝 Create a detailed report and highlight issues in Excel? (y/n): ").strip().lower()
            
            if report_choice in ['y', 'yes']:
                # Create a report with all issues from the session
                create_report_and_highlight(file_path, sheet_name, issues, session_issues)
            else:
                print("Report creation skipped.")
            
            print("\n✅ Analysis complete.")
            
            # Ask what to do next
            print(f"\n🔧 OPTIONS:")
            print("1. Analyze more columns")
            print("2. Exit")
            
            choice = input("\nEnter your choice (1-2): ").strip()
            
            if choice == "1":
                continue
            else:
                break
        else:
            retry = input("\nTry again? (y/n): ").strip().lower()
            if retry not in ['y', 'yes']:
                break
    
    print("\n✨ Done! Thanks for using the Enhanced Excel Validator!")

# Run the main function
if __name__ == "__main__":
    main()